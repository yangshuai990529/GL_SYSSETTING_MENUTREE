#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse the 音画 menu-tree xlsx files into a structured JSON tree."""
import argparse
import json
import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = BASE / "platform" / "data"
DEFAULT_OUTPUT = BASE / "platform" / "data" / "menu_data.json"

LEVEL_HEADER_PAT = re.compile(
    r"(小标题|一级|二级|三级|四级|五级|六级|first\s*level|second\s*level|third\s*level|fourth\s*level|fifth\s*level|sixth\s*level|菜单内容)",
    re.I,
)

def norm(v):
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").strip()

def find_header_row(ws):
    """Header row = the row containing the most level-keywords, scanning first 4 rows."""
    best, best_score = 1, -1
    for r in range(1, 5):
        score = 0
        for c in range(1, ws.max_column + 1):
            if LEVEL_HEADER_PAT.search(norm(ws.cell(r, c).value)):
                score += 1
        if score > best_score:
            best_score, best = score, r
    return best

def classify_columns(ws, hr):
    """Return (level_cols in order, other_cols {col: header}). level_cols are the tree-depth columns."""
    level_cols, other_cols = [], {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(hr, c).value)
        if not h:
            continue
        # Level columns: the depth columns. 菜单内容 (values) is NOT a depth col.
        if LEVEL_HEADER_PAT.search(h) and not re.search(r"菜单内容", h):
            level_cols.append(c)
        else:
            other_cols[c] = h.replace("\n", " ")
    return level_cols, other_cols

def parse_sheet(ws, region, category):
    hr = find_header_row(ws)
    level_cols, other_cols = classify_columns(ws, hr)
    nodes = []
    # stack holds (depth_index, node) for building parent-child
    stack = []
    node_id = 0
    for r in range(hr + 1, ws.max_row + 1):
        # depth = which level column has a value on this row
        depth = None
        label = ""
        for i, c in enumerate(level_cols):
            v = norm(ws.cell(r, c).value)
            if v:
                depth = i
                label = v
                break
        # gather other fields for this row
        fields = {}
        for c, h in other_cols.items():
            v = norm(ws.cell(r, c).value)
            if v:
                fields[h] = v
        row_blank = (depth is None) and (not fields)
        if row_blank:
            continue
        if depth is None:
            # continuation / detail row belonging to the last node: merge fields
            if stack:
                target = stack[-1][1]
                for k, v in fields.items():
                    if k in target["fields"] and v not in target["fields"][k]:
                        target["fields"][k] += "\n---\n" + v
                    elif k not in target["fields"]:
                        target["fields"][k] = v
            continue
        # new node at `depth`
        node_id += 1
        node = {
            "id": f"{region}|{category}|{node_id}",
            "label": label,
            "depth": depth,
            "region": region,
            "category": category,
            "fields": dict(fields),
            "children": [],
        }
        # pop stack until top has depth < current
        while stack and stack[-1][0] >= depth:
            stack.pop()
        if stack:
            node["parent"] = stack[-1][1]["id"]
            stack[-1][1]["children"].append(node)
        else:
            node["parent"] = None
            nodes.append(node)
        stack.append((depth, node))
    return nodes, [other_cols[c] for c in other_cols]

def region_of(filename, sheet):
    if "海外" in filename or "Android" in filename:
        base = "海外"
    else:
        base = "中国区"
    # sub-variants encoded in sheet name
    m = re.search(r"\((X925Pro|NA)\)", sheet)
    if m:
        return f"{base}·{m.group(1)}"
    return base

def category_of(sheet):
    s = sheet
    if "图像" in s or "Picture" in s:
        return "图像"
    if "声音" in s or "Sound" in s:
        return "声音"
    if "Screen" in s:
        return "屏幕设置"
    if "Audio Output" in s:
        return "音频输出"
    return s

def main():
    parser = argparse.ArgumentParser(description="将 MenuTree Excel 文件解析为 menu_data.json")
    parser.add_argument("input_dir", nargs="?", type=Path, default=DEFAULT_INPUT, help="包含 .xlsx 文件的目录，默认 platform/data")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT, help="输出 JSON 路径")
    args = parser.parse_args()
    input_dir = args.input_dir.expanduser().resolve()
    files = sorted(input_dir.glob("*.xlsx"))
    if not files:
        parser.error(f"在 {input_dir} 中未找到 .xlsx 文件")

    result = {"sources": [], "trees": []}
    for path in files:
        fname = path.name
        wb = openpyxl.load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            region = region_of(fname, ws.title)
            category = category_of(ws.title)
            nodes, headers = parse_sheet(ws, region, category)
            def count(ns):
                t = 0
                for n in ns:
                    t += 1 + count(n["children"])
                return t
            total = count(nodes)
            result["trees"].append({
                "file": fname,
                "sheet": ws.title,
                "region": region,
                "category": category,
                "field_headers": headers,
                "roots": nodes,
                "node_count": total,
            })
            print(f"{region:12} / {category:6} [{ws.title:22}] roots={len(nodes):3} nodes={total}")
        result["sources"].append(fname)
    out = args.output.expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as fh:
        json.dump(result, fh, ensure_ascii=False, indent=2)
    print("\nWROTE", out, out.stat().st_size, "bytes")

if __name__ == "__main__":
    main()
